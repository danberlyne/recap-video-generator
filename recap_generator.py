#!/usr/bin/env python3
# recap_generator.py - Extracts clips from video files and combines them into a single recap video.

import openpyxl
from pathlib import Path
from moviepy import VideoFileClip, CompositeVideoClip, ImageClip, TextClip
import moviepy.audio.fx as afx
import moviepy.video.fx as vfx
from moviepy.video.tools.subtitles import SubtitlesClip
from pychorus import find_and_output_chorus
from typing import Literal
import jsonschema
import ruamel.yaml as ry
from input.default import default_schema
import os, sys
from audioread.exceptions import NoBackendError
import warnings

warnings.filterwarnings('ignore')
# sys.tracebacklimit = 0

fschema = os.path.join(os.path.dirname(os.path.realpath(__file__)), "input/schema.yaml")


def generate_recap(# The path to the video data spreadsheet. Must be an XLSX file.
                   video_data_file: str = 'video_data.xlsx',
                   # The path to the directory containing the video files to be used for the recap.
                   video_directory: str = 'Videos',
                   # The location to save the generated recap.
                   output_file: str = 'recap.mp4',
                   # Determines whether to select clips by automatically detecting choruses or by using manually specified start and end times.
                   clip_selection_method: Literal['auto', 'manual'] = 'auto',
                   # The clip length in seconds. Note, this only applies to automatically selected clips. Manually selected clips will not be modified.
                   clip_length: int = 15,
                   # The alignment of the subtitles.
                   sub_alignment: Literal['left', 'center', 'right'] = 'left',
                   # The font file to be used for the subtitle font.
                   sub_font_file: str = 'Fonts/LiberationSans-Regular.ttf',
                   # The font size to be used for the subtitle font.
                   sub_font_size: int = 50,
                   # The colour to be used for the subtitle text. Accepts colour names or hex values.
                   sub_text_color: str = 'white',
                   # The colour to be used for the subtitle text outline. Accepts colour names or hex values.
                   sub_stroke_color: str = 'black',
                   # The thickness of the subtitle text outline.
                   sub_stroke_width: int = 3,
                   # Determines whether the first clip is to be treated as an intro clip.
                   include_intro: bool = False,
                   # If True, overlays a user-specified image over the intro clip.
                   use_overlay_intro_image: bool = False,
                   # The path to the intro image overlay PNG file.
                   intro_image_file: str = 'intro.png',
                   # The intro image will appear over the intro clip for the specified number of seconds.
                   intro_image_duration: int = 10,
                   # If True, resizes the intro image overlay so that it covers the whole screen (while maintaining its aspect ratio).
                   make_intro_image_fullscreen: bool = True,
                   # The font file to be used for the intro font.
                   intro_font_file: str = 'Fonts/LiberationSans-Bold.ttf',
                   # The font size to be used for the intro font.
                   intro_font_size: int = 150,
                   # The colour to be used for the intro text. Accepts colour names or hex values.
                   intro_text_color: str = 'white',
                   # The colour to be used for the intro text outline. Accepts colour names or hex values.
                   intro_stroke_color: str = 'black',
                   # The thickness of the intro text outline.
                   intro_stroke_width: int = 3
                   ) -> None:
    wb = openpyxl.load_workbook(video_data_file)
    ws = wb.active

    print('Importing data from spreadsheet...')
    ids, id_cells = get_video_filenames(ws)

    print('Extracting clips...')
    video_clips = extract_clips(ws, ids, id_cells, clip_selection_method, clip_length, video_directory)

    if use_overlay_intro_image:
        start_clip_idx = 1
    else:
        start_clip_idx = 0
    
    print('Resizing clips...')
    resized_clips = resize_clips(video_clips, start_clip_idx, intro_image_duration, make_intro_image_fullscreen, intro_image_file)

    print('Adding crossfade to clips...')
    custom_padding = 1
    faded_clips = add_crossfade(resized_clips, custom_padding)
    
    print('Concatenating clips...')
    video = CompositeVideoClip(faded_clips)

    print('Generating subtitles...')
    subtitles, start_sub_idx = generate_subtitles(ws, id_cells, video_clips, custom_padding, include_intro, 
                                                  sub_font_file, sub_font_size, sub_text_color, sub_stroke_color, sub_stroke_width,
                                                  intro_font_file, intro_font_size, intro_text_color, intro_stroke_color, intro_stroke_width)

    # Align subtitles according to user input.
    print('Adding subtitles...')
    recap = CompositeVideoClip([video] + 
                               [sub.with_position(('center','center')) for sub in subtitles[:start_sub_idx]] +    # Align intro text in centre
                               [sub.with_position((f'{sub_alignment}','bottom')) for sub in subtitles[start_sub_idx:]])

    # Save recap video file.
    print('Saving recap...')
    recap.write_videofile(output_file)

def get_video_filenames(ws):
    # Get video filenames from spreadsheet.
    id_cells = [cell for cell in ws['A'] if cell.value != 'FILENAME' and cell.value != None and cell.value != '']
    ids = [str(cell.value) for cell in ws['A'] if cell.value != 'FILENAME' and cell.value != None and cell.value != '']
    return ids, id_cells

def extract_clips(ws, ids, id_cells, clip_selection_method, clip_length, video_dir):
    if clip_selection_method == 'manual':
        # Pick out the specified clips from the files and normalise their audio.
        video_clips = []
        for i in range(len(ids)):
            video_clips.append(VideoFileClip(str(Path(video_dir, f'{ids[i]}'))).subclipped(str(ws[f'B{id_cells[i].row}'].value), str(ws[f'C{id_cells[i].row}'].value)).with_effects([afx.AudioNormalize()]))
    elif clip_selection_method == 'auto':
        chorus_error = False
        missing_choruses = []
        video_clips = []
        for i in range(len(ids)):
            print(f'Extracting clip {i+1} of {len(ids)}')
            # Use manual clip if it is specified in spreadsheet.
            if ws[f'B{id_cells[i].row}'].value and ws[f'C{id_cells[i].row}'].value:
                video_clips.append(VideoFileClip(str(Path(video_dir, f'{ids[i]}'))).subclipped(str(ws[f'B{id_cells[i].row}'].value), str(ws[f'C{id_cells[i].row}'].value)).with_effects([afx.AudioNormalize()]))
            # Otherwise select clip automatically via chorus detection.
            else:
                chorus_start = find_and_output_chorus(str(Path(video_dir, f'{ids[i]}')), None)
                try:
                    video_clips.append(VideoFileClip(str(Path(video_dir, f'{ids[i]}'))).subclipped(chorus_start, chorus_start + clip_length).with_effects([afx.AudioNormalize()]))
                except TypeError:
                    chorus_error = True
                    print(f'No chorus found for video {ids[i]}. Please choose a clip manually.')
                    missing_choruses.append(ids[i])
        if chorus_error:
            print(f'Auto-generation failed for some clips. Please choose clips manually for the videos specified below then try again.\n{missing_choruses}')
            input('Press Enter to exit.')
            raise TypeError('Auto-generation failed for some clips.')
    
    return video_clips

def resize_clips(video_clips, start_clip_idx, intro_image_duration, fullscreen_intro_image, intro_image_file):
    # Resize clips while maintaining aspect ratio and then add black borders if necessary to reach 1920x1080p.
    # Also add intro image overlay.
    resized_clips = []

    for clip in video_clips[:start_clip_idx]:
        img_clip = ImageClip(intro_image_file, duration=intro_image_duration)
        if fullscreen_intro_image:
            img_size_ratio = 1.0
        else:
            img_size_ratio = 0.5
        
        if clip.w / clip.h <= 1920 / 1080 and img_clip.w / img_clip.h <= 1920 / 1080:
            resized_clips.append(CompositeVideoClip([ImageClip('1920x1080-black.jpg', duration=clip.duration).resized((1920, 1080)), 
                                                clip.resized(height=1080).with_position('center', 'center'),
                                                img_clip.resized(height=round(1080*img_size_ratio)).with_position('center', 'center')]))
        elif clip.w / clip.h > 1920 / 1080 and img_clip.w / img_clip.h <= 1920 / 1080:
            resized_clips.append(CompositeVideoClip([ImageClip('1920x1080-black.jpg', duration=clip.duration).resized((1920, 1080)), 
                                                clip.resized(width=1920).with_position('center', 'center'),
                                                img_clip.resized(height=round(1080*img_size_ratio)).with_position('center', 'center')])) 
        elif clip.w / clip.h <= 1920 / 1080 and img_clip.w / img_clip.h > 1920 / 1080:
            resized_clips.append(CompositeVideoClip([ImageClip('1920x1080-black.jpg', duration=clip.duration).resized((1920, 1080)), 
                                                clip.resized(height=1080).with_position('center', 'center'),
                                                img_clip.resized(width=round(1920*img_size_ratio)).with_position('center', 'center')])) 
        else: 
            resized_clips.append(CompositeVideoClip([ImageClip('1920x1080-black.jpg', duration=clip.duration), 
                                              clip.resized(width=1920).with_position('center', 'center'),
                                              img_clip.resized(width=round(1920*img_size_ratio)).with_position('center', 'center')]))

    resized_clips += [CompositeVideoClip([ImageClip('1920x1080-black.jpg', duration=clip.duration).resized((1920, 1080)), 
                                         clip.resized(height=1080).with_position('center', 'center')]) 
                      if clip.w / clip.h <= 1920 / 1080 
                      else CompositeVideoClip([ImageClip('1920x1080-black.jpg', duration=clip.duration), 
                                               clip.resized(width=1920).with_position('center', 'center')]) 
                      for clip in video_clips[start_clip_idx:]
                     ]
    
    return resized_clips

def add_crossfade(resized_clips, custom_padding):
    # Add crossfade to clips.
    faded_clips = [resized_clips[0].with_effects([afx.AudioFadeIn(custom_padding), afx.AudioFadeOut(custom_padding), vfx.FadeIn(custom_padding)])]
    idx = resized_clips[0].duration - custom_padding
    for i in range(len(resized_clips[1:])):
        clip = resized_clips[i+1].with_effects([afx.AudioFadeIn(custom_padding), afx.AudioFadeOut(custom_padding)])
        faded_clips.append(clip.with_start(idx).with_effects([vfx.CrossFadeIn(custom_padding)]))
        idx += clip.duration - custom_padding
    # Fade the final clip out to black.
    faded_clips[-1] = faded_clips[-1].with_effects([vfx.FadeOut(custom_padding)])

    return faded_clips

def generate_subtitles(ws, id_cells, video_clips, custom_padding, include_intro, 
                       sub_font_file, sub_font_size, sub_text_color, sub_stroke_color, sub_stroke_width,
                       intro_font_file, intro_font_size, intro_text_color, intro_stroke_color, intro_stroke_width):
    subtitles = []
    start_clip_idx = 0
    start_sub_idx = 0

    if include_intro:
        start_clip_idx = 1
       
        generator = lambda txt: TextClip(text=txt, 
                                         font=intro_font_file, 
                                         font_size=intro_font_size, 
                                         color=intro_text_color, 
                                         stroke_color=intro_stroke_color, 
                                         stroke_width=intro_stroke_width)

        intro_txt = ''
        if ws[f'D{id_cells[0].row}'].value:
            intro_txt += str(ws[f'D{id_cells[0].row}'].value)
        if ws[f'E{id_cells[0].row}'].value:
            intro_txt += '\n' + str(ws[f'E{id_cells[0].row}'].value)
        if ws[f'F{id_cells[0].row}'].value:
            intro_txt += '\n' + str(ws[f'F{id_cells[0].row}'].value)
        
        intro_sub = [
                    ((custom_padding, # Start time of subtitle
                    int(video_clips[0].duration - custom_padding)), # End time of subtitle
                    intro_txt) # Text of subtitle
                    ]
        subtitles.append(SubtitlesClip(intro_sub, make_textclip=generator, encoding='utf-8'))

        start_sub_idx = len(subtitles)

    generator = lambda txt: TextClip(text=txt, 
                                     font=sub_font_file, 
                                     font_size=sub_font_size, 
                                     color=sub_text_color, 
                                     stroke_color=sub_stroke_color, 
                                     stroke_width=sub_stroke_width, 
                                     margin=(10,20))
    subs = []
    for clip in video_clips[start_clip_idx:]:
        sub_txt = ''
        if ws[f'D{id_cells[video_clips.index(clip)].row}'].value:
            sub_txt += str(ws[f'D{id_cells[video_clips.index(clip)].row}'].value)
        if ws[f'E{id_cells[video_clips.index(clip)].row}'].value:
            sub_txt += '\n' + str(ws[f'E{id_cells[video_clips.index(clip)].row}'].value)
        if ws[f'F{id_cells[video_clips.index(clip)].row}'].value:
            sub_txt += '\n' + str(ws[f'F{id_cells[video_clips.index(clip)].row}'].value)
        
        start_time = int(sum(previous_clip.duration - custom_padding for previous_clip in video_clips[:video_clips.index(clip)])) + custom_padding
        end_time = int(sum(previous_clip.duration - custom_padding for previous_clip in video_clips[:video_clips.index(clip)+1]))

        subs.append(((start_time, end_time), sub_txt))

    if subs:
        subtitles.append(SubtitlesClip(subs, make_textclip=generator, encoding='utf-8'))

    return subtitles, start_sub_idx

def read_yaml(finput):
    yaml_schema = load_yaml(fschema) if isinstance(fschema, str) else fschema
    myobj = load_yaml(finput) if isinstance(finput, str) else finput
    DefaultValidatingDraft7Validator = extend_with_default(jsonschema.Draft7Validator)
    DefaultValidatingDraft7Validator(yaml_schema).validate(myobj)
    return myobj

def load_yaml(fname_input):
    reader = ry.YAML(typ="safe", pure=True)
    try:
        with open(fname_input, "r", encoding="utf-8") as f:
            input_yaml = reader.load(f)
    except FileNotFoundError:
        input_yaml = default_schema
    return input_yaml

def extend_with_default(validator_class):
    validate_properties = validator_class.VALIDATORS["properties"]

    def set_defaults(validator, properties, instance, schema):
        for property, subschema in properties.items():
            if "default" in subschema:
                instance.setdefault(property, subschema["default"])

        for error in validate_properties(validator, properties, instance, schema):
            yield error

    return jsonschema.validators.extend(validator_class, {"properties": set_defaults})


if __name__ == '__main__':
    options = read_yaml('options.yaml')
    try:
        generate_recap(**options)
    except NoBackendError:
        raise NoBackendError('No audio backend detected. Please install FFmpeg or another backend.')
    
    try:
        input('Recap generation complete! Press Enter to exit.')
    except EOFError:
        print('\rRecap generation complete!                     ')